"""
Camera Statistics Script
========================

This script provides a utility to manage and process files within a specified directory. The main features include:
- Classifying files based on their extensions.
- Generating statistical summaries.
- Handling image files.

Modules Imported:
-----------------
- os: Provides functions to interact with the operating system.
- pandas: Used for data manipulation and analysis.
- tkinter: Provides classes for creating graphical user interfaces.
- openpyxl.styles: Used for styling Excel files.
- shutil: Offers a number of high-level operations on files and collections of files.
- subprocess: Allows spawning new processes, connecting to their input/output/error pipes, and obtaining their return codes.
- PIL (Pillow): Adds image processing capabilities to Python.

Global Variables:
-----------------
- directorio_actual: Stores the path of the current directory where the script is located.
- carpeta_seleccionada: Global variable to store the selected folder.
- ruta_archivo_estadisticos: Global variable to store the path of the statistical file.

Functions:
----------
- clasificar_archivos(carpeta_principal)
    Classifies files into subfolders based on their extensions.

    Parameters:
    - carpeta_principal (str): The main directory to classify files.

- seleccionar_carpeta_principal()
    Opens a file dialog to select the main folder and calls the classification function.

- buscar_subcarpetas_txt(carpeta)
    Searches for all subfolders named 'TXT' within the specified directory.

    Parameters:
    - carpeta (str): The directory to search for 'TXT' subfolders.

- generar_estadisticos()
    Generates statistics from text files in the selected folder using file dialogs.

- abrir_estadisticos()
    Opens the generated statistics file if available.

- actualizar_etiqueta_ruta()
    Updates the GUI label with the path of the statistics file.

- generar_estadisticos_datos_especificos()
    Opens a new window to specify parameters and generate specific statistics.

- obtener_datos_camara()
    Opens a window to enter IP address for camera connection and file selection.

"""
import os
import threading
import time
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl.styles import PatternFill
from datetime import datetime
import shutil
import subprocess
from PIL import Image, ImageTk

# Obtener la ruta del directorio actual donde se encuentra el script
directorio_actual = os.path.dirname(os.path.abspath(__file__))

# Variable global para almacenar la carpeta seleccionada
carpeta_seleccionada = None
ruta_archivo_estadisticos = None  # Variable global para almacenar la ruta del archivo de estadísticos
idioma = "ES"

# Función para clasificar archivos en subcarpetas según su extensión
def clasificar_archivos(carpeta_principal):
    carpetas = ['png', 'jpg', 'txt']
    # Recorrer la estructura de directorios de la carpeta principal
    for carpeta, _, archivos in os.walk(carpeta_principal):
        for archivo in archivos:
            archivo_path = os.path.join(carpeta, archivo)
            # Obtener la extensión del archivo
            extension = archivo.split('.')[-1].lower()
            # Verificar si la extensión está en la lista de extensiones a clasificar
            if extension in carpetas:
                # Crear la subcarpeta si no existe
                subcarpeta = os.path.join(carpeta, extension.lower())
                if not os.path.exists(subcarpeta):
                    os.makedirs(subcarpeta)
                # Mover el archivo a la subcarpeta correspondiente
                shutil.move(archivo_path, os.path.join(subcarpeta, archivo))

# Función para seleccionar una carpeta principal y clasificar sus archivos
def seleccionar_carpeta_principal():
    carpeta_principal = filedialog.askdirectory()
    if carpeta_principal:
        clasificar_archivos(carpeta_principal)
        if(idioma == "EN"):
            mensaje_label.config(text="Archivos clasificados correctamente.")
        else:
            mensaje_label.config(text="Files classified correctly.")

# Función para buscar todas las subcarpetas llamadas 'TXT'
def buscar_subcarpetas_txt(carpeta):
    subcarpetas_txt = []
    for root, dirs, _ in os.walk(carpeta):
        for dir_name in dirs:
            if dir_name.lower() == 'txt':
                subcarpetas_txt.append(os.path.join(root, dir_name))
    return subcarpetas_txt

# Función para generar estadísticas de archivos de texto en la carpeta seleccionada
def generar_estadisticos():
    global carpeta_seleccionada, ruta_archivo_estadisticos
    # Seleccionar una carpeta
    if (idioma == "EN"):
        carpeta_seleccionada = filedialog.askdirectory(title="Select Folder")
    else:
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar Carpeta")

    if carpeta_seleccionada:
        # Buscar todas las subcarpetas llamadas 'TXT'
        subcarpetas_txt = buscar_subcarpetas_txt(carpeta_seleccionada)
        if subcarpetas_txt:
            # Mostrar mensaje informativo y procesar archivos .txt
            if (idioma == "EN"):
                messagebox.showinfo("Generate Statistics", f"Statistics are generated from the folder: {carpeta_seleccionada}")
            else:
                messagebox.showinfo("Generar Estadísticos",
                                    f"Se generarán estadísticos de la carpeta: {carpeta_seleccionada}")
            valores = []
            for subcarpeta_txt in subcarpetas_txt:
                for root, _, archivos in os.walk(subcarpeta_txt):
                    for archivo in archivos:
                        if archivo.endswith('.txt'):
                            datos = {}
                            ruta_archivo = os.path.join(root, archivo)
                            with open(ruta_archivo, 'r') as file:
                                contenido = file.readlines()
 
                                recipe_id = None
                                exposure_time = None
                                image_time_stamp = None  # Agregar para almacenar el Image Time Stamp
                                blob_data = {}
                                for i in range(1, 7):
                                    blob_key = f'Blob {i}'
                                    blob_data[blob_key] = {
                                        'Enabled': False,
                                        'Threshold': None,
                                        'Min': None,
                                        'Max': None,
                                        'Area': None
                                    }
 
                                for line in contenido:
                                    if "Recipe ID" in line:
                                        recipe_id = int(line.split(': ')[1])
                                    elif "Exposure Time" in line:
                                        exposure_time = int(line.split(': ')[1])
                                    elif "Image Time Stamp" in line:  # Buscar el Image Time Stamp
                                        image_time_stamp = line.split(': ')[1].strip()
                                    elif "Blob" in line and "Enabled: True" in line:
                                        blob_number = int(line.split()[1])
                                        blob_data[f'Blob {blob_number}']['Enabled'] = True
                                    elif "Blob" in line and "Threshold" in line:
                                        blob_number = int(line.split()[1])
                                        blob_data[f'Blob {blob_number}']['Threshold'] = int(line.split(': ')[1])
                                    elif "Blob" in line and "Min" in line:
                                        blob_number = int(line.split()[1])
                                        blob_data[f'Blob {blob_number}']['Min'] = int(line.split(': ')[1])
                                    elif "Blob" in line and "Max" in line:
                                        blob_number = int(line.split()[1])
                                        blob_data[f'Blob {blob_number}']['Max'] = int(line.split(': ')[1])
                                    elif "Blob" in line and "Area" in line:
                                        blob_number = int(line.split()[1])
                                        blob_data[f'Blob {blob_number}']['Area'] = int(line.split(': ')[1])
 
                                if recipe_id is not None and exposure_time is not None and image_time_stamp is not None:
                                    camara = archivo[:19]
                                    datos = {'Camara': camara, 'Archivo': archivo, 'Recipe ID': recipe_id, 'Exposure Time': exposure_time, 'Image Time Stamp': image_time_stamp}
                                    for blob_key, blob_info in blob_data.items():
                                        if blob_info['Enabled']:
                                            datos.update({
                                                f'{blob_key} Threshold': blob_info['Threshold'],
                                                f'{blob_key} Min': blob_info['Min'],
                                                f'{blob_key} Max': blob_info['Max'],
                                                f'{blob_key} Area': blob_info['Area']
                                            })
                                    valores.append(datos)
 
            # Crear un DataFrame de Pandas con los valores de Recipe ID, Exposure Time, Image Time Stamp, Blob 1 Threshold, Blob 1 Min, Blob 1 Max y Blob 1 Area
            df = pd.DataFrame(valores)
            # Obtener el nombre de la carpeta seleccionada
            nombre_carpeta = os.path.basename(carpeta_seleccionada)
 
            # Guardar el DataFrame en un archivo de Excel con el nombre de la carpeta en la raíz de la carpeta seleccionada
            ruta_excel = os.path.join(carpeta_seleccionada, f'{nombre_carpeta}.xlsx')
 
            # Agregar la columna con los nombres de los archivos .txt
            df['Archivo'] = [os.path.basename(path) for path in df['Archivo']]
 
            # Escribir el DataFrame en el archivo Excel con filtro en la primera fila
            with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=True)
                # Ajustar automáticamente el tamaño de las columnas al contenido
                for column in writer.sheets['Sheet1'].columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    writer.sheets['Sheet1'].column_dimensions[column[0].column_letter].width = adjusted_width
 
                # Resaltar en rojo las columnas que tienen campos vacíos
                for col in writer.sheets['Sheet1'].iter_cols():
                    for cell in col:
                        if cell.value is None or cell.value == '':
                            cell.fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
 
            # Almacenar la ruta del archivo generado
            ruta_archivo_estadisticos = ruta_excel
            actualizar_etiqueta_ruta()
            if(idioma=="EN"):
                messagebox.showinfo("Generate Statistics", f"Statistics generated and saved in '{ruta_excel}'")
            else:
                messagebox.showinfo("Generar Estadísticos", f"Estadísticos generados y guardados en '{ruta_excel}'")
        else:
            # Mostrar mensaje de error si no se encuentra ninguna subcarpeta 'TXT'
            if(idioma=="EN"):
                messagebox.showerror("Error", "No 'TXT' subfolders were found in the selected folder.")
            else:
                messagebox.showerror("Error", "No se encontraron subcarpetas 'TXT' en la carpeta seleccionada.")
 
 
# Función para abrir el archivo de estadísticos
def abrir_estadisticos():
    global ruta_archivo_estadisticos
    if ruta_archivo_estadisticos:
        os.startfile(ruta_archivo_estadisticos)
    else:
        if(idioma=="EN"):
            messagebox.showinfo("Open Statistics", "First generate the statistics to open the file.")
        else:
            messagebox.showinfo("Abrir Estadísticos", "Primero genera los estadísticos para abrir el archivo.")
 
# Función para actualizar la etiqueta con la ruta del archivo de estadísticos
def actualizar_etiqueta_ruta():
    if ruta_archivo_estadisticos:
        if(idioma=="EN"):

            etiqueta_ruta.config(text=f"Path of the statistics file: {ruta_archivo_estadisticos}")
        else:
            etiqueta_ruta.config(text=f"Ruta del archivo de estadísticos: {ruta_archivo_estadisticos}")
    else:
        etiqueta_ruta.config(text="")
 
# Función para generar estadísticos de datos específicos
def generar_estadisticos_datos_especificos():
    # Función para procesar los parámetros ingresados y buscar el texto en los archivos .txt
    def procesar_parametros():
        parametros = ["Recipe ID", "Exposure Time", "Image Time Stamp"] + entrada_parametros.get().split(';')

        if(idioma=="EN"):
            carpeta_seleccionada = filedialog.askdirectory(title="Select Folder")
        else:
            carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar Carpeta")
 
        if not parametros:
            if(idioma=="EN"):
                messagebox.showerror("Error", "Please enter at least one parameter.")
                return
            else:
                messagebox.showerror("Error", "Por favor ingresa al menos un parámetro.")
                return
 
        if not carpeta_seleccionada:
            if(idioma=="EN"):
                messagebox.showerror("Error", "You must select a folder.")
                return
            else:
                messagebox.showerror("Error", "Debes seleccionar una carpeta.")
                return

        archivos_encontrados = []
        nombres_archivos = []  # Lista para almacenar los nombres de los archivos .txt encontrados
        camaras_encontradas =[]
        for root, _, archivos in os.walk(carpeta_seleccionada):
            for archivo in archivos:
                if archivo.endswith('.txt'):
                    ruta_archivo = os.path.join(root, archivo)
                    with open(ruta_archivo, 'r') as file:
                        contenido = file.read()
                        if all(parametro in contenido for parametro in parametros):
                            archivos_encontrados.append(ruta_archivo)
                            nombres_archivos.append(archivo)
                            camaras_encontradas.append(archivo[:19])
 
        if not archivos_encontrados:
            if(idioma=="EN"):
                messagebox.showinfo("Result", "No matches were found.")
                return
            else:
                messagebox.showinfo("Resultado", "No se encontraron coincidencias.")
                return
 
        generar_excel(archivos_encontrados, nombres_archivos, parametros, carpeta_seleccionada, camaras_encontradas)
 
    # Función para generar un archivo Excel con los datos de los archivos encontrados
    def generar_excel(archivos, nombres_archivos, parametros, carpeta_seleccionada, camaras_encontradas):
                             
        datos = {'Camaras': camaras_encontradas, 'Archivo': nombres_archivos}  # Inicializar el diccionario con los nombres de los archivos .txt
        for parametro in parametros:
            datos[parametro] = []
 
        for archivo in archivos:
            with open(archivo, 'r') as file:
                contenido = file.readlines()
                for linea in contenido:
                    for parametro in parametros:
                        if parametro in linea:
                            valor = linea.split(': ')[1].strip()
                            datos[parametro].append(valor)

        nombre_carpeta = os.path.basename(carpeta_seleccionada)
        df = pd.DataFrame(datos)
        ruta_excel = os.path.join(carpeta_seleccionada, f'{nombre_carpeta}_datos_especificos.xlsx')
        df.to_excel(ruta_excel, index=False, header=True)

        for column in df.columns:
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
            df[column] = df[column].apply(lambda x: x.ljust(max_length))

        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=True)
            for column in writer.sheets['Sheet1'].columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                writer.sheets['Sheet1'].column_dimensions[column[0].column_letter].width = adjusted_width

        if(idioma=="EN"):
            messagebox.showinfo("Result", f"Statistics generated and saved in '{ruta_excel}'")
        else:
            messagebox.showinfo("Resultado", f"Estadísticos generados y guardados en '{ruta_excel}'")

    ventana_parametros = tk.Toplevel()
    if(idioma=="EN"):
        ventana_parametros.title("Generate Specific Data Statistics")
    else:
        ventana_parametros.title("Generar Estadísticos de Datos Específicos")

    ventana_parametros.geometry("400x200")

    if (idioma == "EN"):
        etiqueta_parametros = ttk.Label(ventana_parametros, text="Parameters (separated by ';'):")
        etiqueta_parametros.pack(pady=10)

        entrada_parametros = ttk.Entry(ventana_parametros, width=50)
        entrada_parametros.pack(pady=10)

        boton_procesar = ttk.Button(ventana_parametros, text="Process", command=procesar_parametros)
        boton_procesar.pack(pady=10)
    else:
        etiqueta_parametros = ttk.Label(ventana_parametros, text="Parámetros (separados por ';'):")
        etiqueta_parametros.pack(pady=10)

        entrada_parametros = ttk.Entry(ventana_parametros, width=50)
        entrada_parametros.pack(pady=10)

        boton_procesar = ttk.Button(ventana_parametros, text="Procesar", command=procesar_parametros)
        boton_procesar.pack(pady=10)


def obtener_datos_camara():
    global ventana_archivos, conjunto_ip, ventana_estado, monitor_conexion, conjunto_estacion
    ventana_estado = None
    ventana_archivos = None
    conjunto_ip = []
    conjunto_estacion = []
    monitor_running = False
    monitor_conexion = None
    direccion_ip_global = None

    def procesar_direccion_ip(direccion_ip=None):
        global direccion_ip_global, monitor_conexion, monitor_running, conjunto_ip, conjunto_estacion

        if not direccion_ip:
            direccion_ip = entrada_ip.get()
            conjunto_ip = []
            conjunto_estacion = []
            if not conjunto_ip:
                conjunto_ip.append(direccion_ip)
            else:
                conjunto_ip[0] = direccion_ip

        if direccion_ip:
            ping_exit_code = subprocess.call(['ping', '-n', '1', direccion_ip], stdout=subprocess.DEVNULL)
            if ping_exit_code == 0:
                try:
                    comando = f"net use \\\\{direccion_ip}\\IPC$ /user:NAM\\mtxuser Matrox"
                    subprocess.run(comando, shell=True, check=True)
                    if(idioma=="EN"):
                        print(f"Conexión SMB establecida con {direccion_ip}")
                    else:
                        print(f"SMB connection established with {direccion_ip}")

                    if ventana_archivos is None:
                        abrir_ventana_seleccion_archivos()

                    if ventana_estado is None:
                        mostrar_ventana_estado(direccion_ip)

                    monitor_running = True
                    if monitor_conexion is None or not monitor_conexion.is_alive():
                        monitor_conexion = threading.Thread(target=monitorizar_conexion, args=(direccion_ip,))
                        monitor_conexion.start()

                except subprocess.CalledProcessError as e:
                    if(idioma=="EN"):
                        messagebox.showerror("Error", f"Failed to establish SMB connection to {direccion_ip}: {str(e)}")
                    else:
                        messagebox.showerror("Error",
                                             f"No se pudo establecer la conexión SMB con {direccion_ip}: {str(e)}")
                    abrir_ventana_credenciales(direccion_ip)
            else:
                if(idioma=="EN"):
                    messagebox.showerror("Error", f"Could not ping IP address {direccion_ip}.")
                else:
                    messagebox.showerror("Error", f"No se pudo hacer ping a la dirección IP {direccion_ip}.")
        else:
            if(idioma=="EN"):
                messagebox.showerror("Error", "You must enter a valid IP address.")
            else:
                messagebox.showerror("Error", "Debes ingresar una dirección IP válida.")
    def mostrar_ventana_estado(direccion_ip):
        global ventana_estado, direccion_ip_global

        if ventana_estado is None:
            ventana_estado = tk.Toplevel(root)
            if(idioma=="EN"):
                ventana_estado.title("Connection Status")
            else:
                ventana_estado.title("Estado de Conexión")

            ventana_estado.geometry("300x150")
            ventana_estado.resizable(False, False)

            if(idioma=="EN"):
                estado_label = ttk.Label(ventana_estado, text=f"IP Address: {direccion_ip}\nStatus: Connected")
            else:
                estado_label = ttk.Label(ventana_estado, text=f"Dirección IP: {direccion_ip}\nEstado: Conectado")
            estado_label.pack(pady=20)

            def cerrar_conexion():
                global monitor_running, ventana_estado
                monitor_running = False
                comando_desconectar = f"net use \\\\{direccion_ip}\\IPC$ /delete"
                subprocess.run(comando_desconectar, shell=True, check=True)
                if(idioma=="EN"):
                    print(f"SMB connection closed with {direccion_ip}")
                else:
                    print(f"Conexión SMB cerrada con {direccion_ip}")

                ventana_estado.destroy()

                if(idioma=="EN"):
                    messagebox.showinfo("Connection Closed", f"Closed connection with {direccion_ip}.")
                else:
                    messagebox.showinfo("Conexión Cerrada", f"Conexión cerrada con {direccion_ip}.")

            def on_closing():
                cerrar_conexion()
                direccion_ip = None
                direccion_ip_global = None
                directorio_actual = None
                conjunto_ip = []

                if(idioma=="EN"):
                    messagebox.showinfo("Connection Lost", f"The connection to {direccion_ip} has been lost.")
                else:
                    messagebox.showinfo("Conexion Perdida", f"Se ha perdido la conexión con {direccion_ip}.")

            boton_cerrar = ttk.Button(ventana_estado, text="Cerrar Conexión", command=cerrar_conexion)
            if(idioma=="EN"):
                boton_cerrar.config(text="Close connection")
            else:
                boton_cerrar.config(text="Cerrar conexión")
            boton_cerrar.pack(pady=10)

            ventana_estado.protocol("WM_DELETE_WINDOW", on_closing)
        else:
            if(idioma=="EN"):
                estado_label.config(text=f"IP Address: {direccion_ip}\nStatus: Connected")
            else:
                estado_label.config(text=f"Dirección IP: {direccion_ip}\nEstado: Conectado")

    def monitorizar_conexion(direccion_ip):
        global direccion_ip_global, ventana_estado

        while monitor_running:
            ping_exit_code = subprocess.call(['ping', '-n', '1', direccion_ip], stdout=subprocess.DEVNULL)
            if ping_exit_code != 0:
                if direccion_ip == direccion_ip_global:
                    if(idioma=="EN"):
                        estado_label.config(text=f"IP Address: {direccion_ip}\nStatus: Disconnected")
                        messagebox.showwarning("Connection Lost", f"The connection to {direccion_ip} has been lost.")
                    else:
                        estado_label.config(text=f"Dirección IP: {direccion_ip}\nEstado: Desconectado")
                        messagebox.showwarning("Conexión Perdida", f"Se ha perdido la conexión con {direccion_ip}.")

                    ventana_estado.destroy()
                    break

            time.sleep(5)

    def abrir_ventana_seleccion_archivos():
        global ventana_archivos, progress_bar, progress_label, ventana_archivos
        ventana_archivos = tk.Toplevel(root)
        if(idioma=="EN"):
            ventana_archivos.title("Select the files to copy:")
        else:
            ventana_archivos.title("Seleccionar Archivos a copiar")
        ventana_archivos.geometry("300x350")
        ventana_archivos.resizable(False, False)
        # Configurar la ventana para que siempre se muestre al frente
        ventana_archivos.attributes('-topmost', True)

        if(idioma=="EN"):
            etiqueta_instrucciones = ttk.Label(ventana_archivos, text="Select the files to copy:")
        else:
            etiqueta_instrucciones = ttk.Label(ventana_archivos, text="Selecciona los archivos a copiar:")
        etiqueta_instrucciones.pack(pady=10)

        var_jpg = tk.IntVar()
        check_jpg = ttk.Checkbutton(ventana_archivos, text="JPG", variable=var_jpg)
        check_jpg.pack()

        var_png = tk.IntVar()
        check_png = ttk.Checkbutton(ventana_archivos, text="PNG", variable=var_png)
        check_png.pack()

        var_txt = tk.IntVar()
        check_txt = ttk.Checkbutton(ventana_archivos, text="TXT", variable=var_txt)
        check_txt.pack()

        # Instrucciones combo box
        if(idioma=="EN"):
            etiqueta_combo = ttk.Label(ventana_archivos, text="Select the inspection result to extract:")
        else:
            etiqueta_combo = ttk.Label(ventana_archivos, text="Selecciona el resultado de la inspección a extraer:")
        etiqueta_combo.pack(pady=10)

        # ComboBox para seleccionar inspección
        if(idioma=="EN"):
            combo_inspeccion = ttk.Combobox(ventana_archivos, values=["Pass", "Fail", "All inspections"])
        else:
            combo_inspeccion = ttk.Combobox(ventana_archivos, values=["Pass", "Fail", "Todas las inspecciones"])

        combo_inspeccion.pack(pady=10)

        # Widget de progreso inicialmente oculto
        if(idioma=="EN"):
            progress_label = ttk.Label(ventana_archivos, text="Extraction Progress:")
        else:
            progress_label = ttk.Label(ventana_archivos, text="Progreso de Extracción:")
        progress_label.pack(pady=10)
        progress_label.pack_forget()
        progress_bar = ttk.Progressbar(ventana_archivos, orient='horizontal', length=100, mode='determinate')
        progress_bar.pack(pady=10)
        progress_bar.pack_forget()  # Ocultar inicialmente el widget de progress

        if(idioma=="EN"):
            boton_extraer = ttk.Button(ventana_archivos, text="Extract Files",
                                   command=lambda: extraer_archivos(var_jpg, var_png, var_txt, combo_inspeccion.get()))
        else:
            boton_extraer = ttk.Button(ventana_archivos, text="Extraer Archivos",
                                       command=lambda: extraer_archivos(var_jpg, var_png, var_txt,
                                                                        combo_inspeccion.get()))
        boton_extraer.pack(pady=10)
        if(idioma=="EN"):
            boton_cerrar = ttk.Button(ventana_archivos, text="Close", command=ventana_archivos.destroy)
            boton_cerrar.pack(pady=10)
        else:
            boton_cerrar = ttk.Button(ventana_archivos, text="Cerrar", command=ventana_archivos.destroy)
            boton_cerrar.pack(pady=10)

    def extraer_archivos(var_jpg, var_png, var_txt, inspeccion=None):
        global conjunto_ip, progess_bar, progress_label, ventana_archivos
        estacion = None

        # Condicion para cuando se selecciona Todas las inspecciones

        if inspeccion == "Todas las inspecciones" or inspeccion == "All inspections":
            inspeccion = None

        ventana_archivos.attributes('-topmost', False)

        if(idioma=="EN"):
            carpeta_destino_padre = filedialog.askdirectory(title="Select the destination folder")
        else:
            carpeta_destino_padre = filedialog.askdirectory(title="Selecciona la carpeta de destino")

        ventana_archivos.attributes('-topmost', True)
        if not carpeta_destino_padre:
            return

        extensiones_seleccionadas = []
        if var_jpg.get():
            extensiones_seleccionadas.append(".jpg")
        if var_png.get():
            extensiones_seleccionadas.append(".png")
        if var_txt.get():
            extensiones_seleccionadas.append(".txt")

        progress_label.pack()
        progress_bar['value'] = 0
        progress_bar.pack()  # Mostrar el widget de progreso

        if not extensiones_seleccionadas:
            if(idioma=="EN"):
                messagebox.showwarning("Advertencia", "No se ha seleccionado ningún tipo de archivo para copiar.")
                return
            else:
                messagebox.showwarning("Warning", "No file type has been selected to copy.")
                return

        for ip in conjunto_ip:
            i = 0
            if conjunto_estacion != []:
                estacion = conjunto_estacion[i]
            else:
                estacion = "Estacion"

            ruta_origen = f"\\\\{ip}\\mtxuser"

            # Crear nombre de la carpeta con "valor en columna"m "-", "dirección IP", "fecha"
            now = datetime.now()
            fecha = now.strftime("%Y-%m-%d_%H-%M-%S")
            nombre_carpeta = f"{estacion}-{ip}-{fecha}"

            # Ruta completa de la carpeta destino
            carpeta_destino = os.path.join(carpeta_destino_padre, nombre_carpeta)

            # Crear la carpeta si no existe
            if not os.path.exists(carpeta_destino):
                os.makedirs(carpeta_destino)

            total_archivos = sum(len(files) for _, _, files in os.walk(ruta_origen))
            progress_bar['maximum'] = total_archivos

            # Copiar archivos según la inspección seleccionada
            for raiz, dirs, archivos in os.walk(ruta_origen):
                for archivo in archivos:
                    if any(archivo.lower().endswith(ext) for ext in extensiones_seleccionadas):
                        if inspeccion is None or inspeccion.lower() in archivo.lower():
                            ruta_completa_origen = os.path.join(raiz, archivo)
                            ruta_completa_destino = os.path.join(carpeta_destino, archivo)
                            try:
                                shutil.copy2(ruta_completa_origen, ruta_completa_destino)
                                if(idioma=="EN"):
                                    print(f"File copied: {ruta_completa_origen} -> {ruta_completa_destino}")
                                else:
                                    print(f"Archivo copiado: {ruta_completa_origen} -> {ruta_completa_destino}")
                                progress_bar['value'] += 1
                                ventana_archivos.update_idletasks()  # Actualizar la ventana para mostrar el progreso

                            except Exception as e:
                                if(idioma=="EN"):
                                    print(f"Error copying file {ruta_completa_origen}: {str(e)}")
                                else:
                                    print(f"Error al copiar el archivo {ruta_completa_origen}: {str(e)}")
            i = i + 1
        if(idioma=="EN"):
            messagebox.showinfo("Complete Extraction", "Files extracted successfully.")
        else:
            messagebox.showinfo("Extracción Completa", "Archivos extraídos correctamente.")
        progress_bar.pack_forget()  # Ocultar el widget de progreso al finalizar

    def abrir_ventana_credenciales(direccion_ip):
        ventana_credenciales = tk.Toplevel(root)
        if(idioma=="EN"):
            ventana_credenciales.title(f"Login Credentials for {direccion_ip}")
        else:
            ventana_credenciales.title(f"Ingresar Credenciales para {direccion_ip}")
        ventana_credenciales.geometry("300x150")
        ventana_credenciales.resizable(False, False)
        if(idioma=="EN"):
            etiqueta_usuario = ttk.Label(ventana_credenciales, text="User:")
            etiqueta_usuario.pack(pady=10)
        else:
            etiqueta_usuario = ttk.Label(ventana_credenciales, text="Usuario:")
            etiqueta_usuario.pack(pady=10)

        entrada_usuario = ttk.Entry(ventana_credenciales)
        entrada_usuario.pack()

        if(idioma=="EN"):
            etiqueta_contrasena = ttk.Label(ventana_credenciales, text="Password:")
            etiqueta_contrasena.pack(pady=10)
        else:
            etiqueta_contrasena = ttk.Label(ventana_credenciales, text="Contraseña:")
            etiqueta_contrasena.pack(pady=10)

        entrada_contrasena = ttk.Entry(ventana_credenciales, show="*")
        entrada_contrasena.pack()

        def guardar_credenciales():
            global usuario_global, contrasena_global
            usuario_global = entrada_usuario.get()
            contrasena_global = entrada_contrasena.get()
            ventana_credenciales.destroy()
            autenticar_conexion(direccion_ip, usuario_global, contrasena_global)

        if(idioma=="EN"):
            boton_guardar = ttk.Button(ventana_credenciales, text="Save", command=guardar_credenciales)
            boton_guardar.pack(pady=20)
        else:
            boton_guardar = ttk.Button(ventana_credenciales, text="Guardar", command=guardar_credenciales)
            boton_guardar.pack(pady=20)

    def autenticar_conexion(direccion_ip, usuario, contrasena):
        try:
            comando = f"net use \\\\{direccion_ip}\\IPC$ /user:{usuario} {contrasena}"
            subprocess.run(comando, shell=True, check=True)
            if(idioma=="EN"):
                print(f"SMB connection established with {direccion_ip}")
            else:
                print(f"Conexión SMB establecida con {direccion_ip}")

            if ventana_archivos is None:
                abrir_ventana_seleccion_archivos()

            if ventana_estado is None:
                mostrar_ventana_estado(direccion_ip)

            monitor_running = True
            if monitor_conexion is None or not monitor_conexion.is_alive():
                monitor_conexion = threading.Thread(target=monitorizar_conexion, args=(direccion_ip,))
                monitor_conexion.start()

        except subprocess.CalledProcessError as e:
            if(idioma=="EN"):
                messagebox.showerror("Error", f"Failed to establish SMB connection to {direccion_ip}: {str(e)}")
            else:
                messagebox.showerror("Error", f"No se pudo establecer la conexión SMB con {direccion_ip}: {str(e)}")
            abrir_ventana_credenciales(direccion_ip)

    def mostrar_ventana_ip():
        ventana_ip = tk.Toplevel(root)
        if(idioma=="EN"):
            ventana_ip.title("Enter IP Address")
        else:
            ventana_ip.title("Ingrese la Dirección IP")

        ventana_ip.geometry("300x150")
        ventana_ip.resizable(False, False)

        if(idioma=="EN"):
            etiqueta_ip = ttk.Label(ventana_ip, text="IP Address:")
            etiqueta_ip.pack(pady=10)
        else:
            etiqueta_ip = ttk.Label(ventana_ip, text="Dirección IP:")
            etiqueta_ip.pack(pady=10)

        global entrada_ip
        entrada_ip = ttk.Entry(ventana_ip)
        entrada_ip.pack()

        if(idioma=="EN"):
            boton_procesar = ttk.Button(ventana_ip, text="Process", command=procesar_direccion_ip)
            boton_procesar.pack(pady=10)
        else:
            boton_procesar = ttk.Button(ventana_ip, text="Procesar", command=procesar_direccion_ip)
            boton_procesar.pack(pady=10)


    def extraer_ips_desde_excel():
        global conjunto_ip, conjunto_estacion
        if(idioma=="EN"):
            archivo_excel = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        else:
            archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx;*.xls")])

        if not archivo_excel:
            return

        df = pd.read_excel(archivo_excel)

        if 'Estacion' not in df.columns:
            if(idioma=="EN"):
                messagebox.showerror("Error", "The Excel file does not contain a column 'Estacion'.")
                return
            else:
                messagebox.showerror("Error", "El archivo de Excel no contiene una columna 'Estacion'.")
                return

        if 'IP' not in df.columns:
            if(idioma=="EN"):
                messagebox.showerror("Error", "The Excel file does not contain a column 'IP'.")
                return
            else:
                messagebox.showerror("Error", "El archivo de Excel no contiene una columna 'IP'.")
                return

        estaciones_disponibles = df['Estacion'].dropna().tolist()
        ips_disponibles = df['IP'].dropna().tolist()

        # Crear un diccionario de IP a estación
        ip_estacion_dict = dict(zip(ips_disponibles, estaciones_disponibles))

        # Crear ventana para seleccionar IPs
        seleccionar_ips = tk.Toplevel(root)
        if(idioma=="EN"):
            seleccionar_ips.title("Select IPs")
        else:
            seleccionar_ips.title("Seleccionar IPs")
        seleccionar_ips.geometry("400x300")
        seleccionar_ips.resizable(False, False)

        if(idioma=="EN"):
            ttk.Label(seleccionar_ips, text="Select the IPs to be processed:").pack(pady=10)
        else:
            ttk.Label(seleccionar_ips, text="Seleccione las IPs a procesar:").pack(pady=10)

        selected_ips = []
        checkboxes = []

        def toggle_ip(ip):
            if ip in selected_ips:
                selected_ips.remove(ip)
            else:
                selected_ips.append(ip)

        # Mostrar checkboxes para cada IP
        for ip in ips_disponibles:
            var_ip = tk.IntVar()
            checkbox = ttk.Checkbutton(seleccionar_ips, text=ip, variable=var_ip, command=lambda ip=ip: toggle_ip(ip))
            checkbox.var = var_ip  # Guardar referencia a la variable IntVar
            checkbox.pack()
            checkboxes.append(checkbox)

        def seleccionar_todo():
            for checkbox in checkboxes:
                checkbox.var.set(1)
                if checkbox.cget("text") not in selected_ips:
                    selected_ips.append(checkbox.cget("text"))
                boton_seleccionar.pack_forget()
                boton_deseleccionar.pack(pady=10)

        def deseleccionar_todo():
            for checkbox in checkboxes:
                checkbox.var.set(0)
                if checkbox.cget("text") not in selected_ips:
                    selected_ips.append(checkbox.cget("text"))
                boton_deseleccionar.pack_forget()
                boton_seleccionar.pack(pady=10)

        def procesar_seleccion():
            global conjunto_ip, conjunto_estacion
            conjunto_ip = selected_ips
            conjunto_estacion = [ip_estacion_dict[ip] for ip in conjunto_ip]
            seleccionar_ips.destroy()

            if(idioma=="EN"):
                messagebox.showinfo("Complete Selection",
                                f"Selected IPs: {conjunto_ip}\nSelected stations: {conjunto_estacion}")
            else:
                messagebox.showinfo("Selección Completa",
                                    f"IPs seleccionadas: {conjunto_ip}\nEstaciones seleccionadas: {conjunto_estacion}")
            for ip in conjunto_ip:
                procesar_direccion_ip(ip)

        if(idioma=="EN"):
            ttk.Button(seleccionar_ips, text="Accept", command=procesar_seleccion).pack(pady=10)

            boton_deseleccionar = ttk.Button(seleccionar_ips, text="Deselect All", command=deseleccionar_todo)
            boton_deseleccionar.pack_forget()
            boton_seleccionar = ttk.Button(seleccionar_ips, text="Select All", command=seleccionar_todo)
            boton_seleccionar.pack(pady=10)
        else:
            ttk.Button(seleccionar_ips, text="Aceptar", command=procesar_seleccion).pack(pady=10)

            boton_deseleccionar = ttk.Button(seleccionar_ips, text="Deseleccionar Todo", command=deseleccionar_todo)
            boton_deseleccionar.pack_forget()
            boton_seleccionar = ttk.Button(seleccionar_ips, text="Seleccionar Todo", command=seleccionar_todo)
            boton_seleccionar.pack(pady=10)

    def mostrar_ventana_archivo():
        ventana_archivo = tk.Toplevel(root)
        if(idioma=="EN"):
            ventana_archivo.title("Archivo")
        else:
            ventana_archivo.title("File")

        # Botón para "Ingresar IP"
        btn_ingresar_ip = ttk.Button(ventana_archivo, text="Ingresar IP", command=mostrar_ventana_ip)
        btn_ingresar_ip.pack(pady=10)

        # Botón para "Extraer direcciones IP desde un excel"
        btn_extraer_ips = ttk.Button(ventana_archivo, text="Extraer IP desde un excel",
                                     command=extraer_ips_desde_excel)
        btn_extraer_ips.pack(pady=10)

        # Separador
        separador = ttk.Separator(ventana_archivo, orient='horizontal')
        separador.pack(fill='x', pady=10)

        # Botón para "Salir"
        btn_salir = ttk.Button(ventana_archivo, text="Salir", command=root.quit)
        btn_salir.pack(pady=10)
        if(idioma=="EN"):
            btn_ingresar_ip.config(text="Enter IP Address")
            btn_extraer_ips.config(text="Extract IP from an Excel file")

    mostrar_ventana_archivo()

# Función para cambiar el idioma de la aplicación
def cambiar_idioma():
    global idioma
    if idioma == 'EN':
        idioma = 'ES'
    else:
        idioma = 'EN'
    actualizar_texto_elementos()

# Función para actualizar el texto de los elementos de la interfaz según el idioma seleccionado
def actualizar_texto_elementos():
    if idioma == 'EN':
        root.title("Data Statistics Generator")
        seleccionar_button.config(text="Classify Files")
        estadisticos_button.config(text="Generate Statistics")
        abrir_estadisticos_button.config(text="Open Statistics")
        estadisticos_especificos_button.config(text="Specific Data Statistics")
        boton_cambiar_idioma.config(text="Español")
        mensaje_label.config(text="Select a main folder to classify files.")
        encabezado_label.config(text="—————————————| FUNCTIONS |——————————————")
        encabezadoSPECIAL_label.config(text="—————————| SPECIAL FUNCTIONS |————————————")
        instrucciones_label.config(text="———————————| INSTRUCTIONS |—————————————\n\n1. Select the folder to select the files to be classified.\n2. Depending on the data of interest, click on 'Generate\n    Stadistics' or if you need any other information click on the button\n    'Specific Data Statistics'.")
        boton_obtener_datos.config(text="Obtaining data from the camera")
        encabezadoIdioma_label.config(text="Change language:")
    else:
        root.title("Generador de Estadísticas de Datos")
        seleccionar_button.config(text="Clasificar Archivos")
        estadisticos_button.config(text="Generar Estadísticos")
        abrir_estadisticos_button.config(text="Abrir Estadísticos")
        estadisticos_especificos_button.config(text="Estadísticos Datos Específicos")
        boton_cambiar_idioma.config(text="English")
        mensaje_label.config(text="Selecciona una carpeta principal para clasificar archivos.")
        encabezado_label.config(text="—————————————| FUNCIONES |——————————————")
        encabezadoSPECIAL_label.config(text="—————————| FUNCIONES ESPECIALES |————————————")
        instrucciones_label.config(text="———————————| INSTRUCCIONES |—————————————\n\n1. Seleccionar la carpeta para seleccionar los archivos a clasificar.\n2. Dependiendo de los datos de interes, dar click en el boton de 'Generar\n    Estadistico' o si se necesita algún otro dato dar click en el botón\n    'Estadisticos Datos Especificos'.")
        boton_obtener_datos.config(text="Obtener datos de la cámara")
        encabezadoIdioma_label.config(text="Cambiar idioma:")

# Configuración de la ventana principal
root = tk.Tk()
root.title("Inspection Tools Statistics V1.01")
root.geometry("450x550")  # Establecer el tamaño inicial de la ventana
root.resizable(False, False)  # Evitar que la ventana se redimensione

# Estilo de Material Design
style = ttk.Style()
style.theme_use("clam")  # Elige el estilo Material Design
style.configure("TLabel", foreground="black", background="#f0f0f0", font=('Roboto', 10))
style.configure("TButton", padding=10, relief="flat", background="#3f51b5", foreground="white", font=('Roboto', 8, 'bold'), width=30)
style.map("TButton", background=[('active', '#283593')])

# Nuevo estilo para los botones en la ventana principal con letra más pequeña
style.configure("BotonesVentanaPrincipal.TButton", font=('Roboto', 7, 'bold'))

# Etiqueta para mostrar las instrucciones
instrucciones_label = ttk.Label(root, text="———————————| INSTRUCCIONES |—————————————\n\n1. Seleccionar la carpeta para seleccionar los archivos a clasificar.\n2. Dependiendo de los datos de interes, dar click en el boton de 'Generar\n    Estadistico' o si se necesita algún otro dato dar click en el botón\n    'Estadisticos Datos Especificos'.")
instrucciones_label.place(x=10, y=380)

# Combinar la ruta del directorio actual con el nombre de la imagen
ruta_imagen_matrox = os.path.join(directorio_actual, "matrox.png")
# Cargar imagen Matrox en la ventana principal
imagen_matrox = tk.PhotoImage(file=ruta_imagen_matrox)

# Widget para mostrar la imagen Matrox en la ventana principal
matrox_label = tk.Label(root, image=imagen_matrox)
matrox_label.place(x=390, y=5)

# Combinar la ruta del directorio actual con el nombre de la imagen
ruta_imagen_vista = os.path.join(directorio_actual, "vista.png")
# Cargar imagen Vista en la ventana principal
imagen_vista = tk.PhotoImage(file=ruta_imagen_vista)

# Widget para mostrar la imagen Vista en la ventana principal
imagen_label = tk.Label(root, image=imagen_vista)
imagen_label.pack(side="top", padx=20, pady=15, anchor="nw")

# Etiqueta para mostrar las encabezado de los botones
encabezado_label = ttk.Label(root, text="—————————————| FUNCIONES |——————————————")
encabezado_label.pack(pady=10, anchor="w", padx=15)

# Botón para generar estadísticos
estadisticos_button = ttk.Button(root, text="Generar Estadísticos Blobs", command=generar_estadisticos, style="TButton")
estadisticos_button.place(x=230, y=107)

# Botón para clasificar
seleccionar_button = ttk.Button(root, text="Clasificar Extensiones", command=seleccionar_carpeta_principal, style="TButton")
seleccionar_button.pack(pady=10, anchor="w", padx=15)

# Botón para abrir el archivo de estadísticos
abrir_estadisticos_button = ttk.Button(root, text="Abrir Estadísticos", command=abrir_estadisticos, style="TButton")
abrir_estadisticos_button.pack(pady=5, anchor="w", padx=115)

# Etiqueta para mostrar las encabezado de los botones
encabezadoSPECIAL_label = ttk.Label(root, text="—————————| FUNCIONES ESPECIALES |————————————")
encabezadoSPECIAL_label.pack(pady=10, anchor="w", padx=15)

# Botón para generar estadísticos de datos específicos
estadisticos_especificos_button = ttk.Button(root, text="Estadísticos Datos Específicos", command=generar_estadisticos_datos_especificos, style="TButton")
estadisticos_especificos_button.pack(pady=5, anchor="w", padx=115)

# Etiqueta para mostrar mensajes
mensaje_label = ttk.Label(root, text="", style="TLabel")
mensaje_label.pack()

# Etiqueta para mostrar la ruta del archivo de estadísticos
etiqueta_ruta = ttk.Label(root, text="", style="TLabel")
etiqueta_ruta.place(x=1000, y=1000)

# Botón para obtener datos de la cámara con estilo personalizado
boton_obtener_datos = ttk.Button(root, text="Obtener datos de la cámara", command=obtener_datos_camara, style="TButton")
boton_obtener_datos.pack(pady=5, anchor="w", padx=115)

# Etiqueta para mostrar las encabezado del boton cambiar idioma
encabezadoIdioma_label = ttk.Label(root, text="Cambiar idioma:")
encabezadoIdioma_label.place(relx=1.0, rely=1.0, x=-100, y=-25, anchor="se")


# Botón para cambiar el idioma de la aplicación
boton_cambiar_idioma = ttk.Button(root, text="English", command=cambiar_idioma, style="TButton", width=10)
boton_cambiar_idioma.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")


# Mostrar la ventana principal
root.mainloop()
